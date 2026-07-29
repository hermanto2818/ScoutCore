from tkinter import filedialog, messagebox
from openpyxl import load_workbook

from database import kosongkan_data_siswa, simpan_siswa


def import_excel():

    file_path = filedialog.askopenfilename(
        title="Pilih File Excel",
        filetypes=[("Excel Files", "*.xlsx *.xls")]
    )

    if not file_path:
        return

    try:

        wb = load_workbook(file_path, data_only=True)
        ws = wb["AWAL JULI"]

        kelas_aktif = None
        data_siswa = []

        daftar_kelas = [
            "KELAS 4 A",
            "KELAS 4 B",
            "KELAS 5 A",
            "KELAS 5 B",
            "KELAS 6 A",
            "KELAS 6 B"
        ]

        for row in ws.iter_rows(values_only=True):

            kolom1 = str(row[0]).strip().upper() if row[0] else ""

            if kolom1 in daftar_kelas:
                kelas_aktif = kolom1.replace("KELAS ", "")
                continue

            if kolom1.startswith("KELAS") and kolom1 not in daftar_kelas:
                kelas_aktif = None
                continue

            if kelas_aktif is None:
                continue

            no = row[0]

            if not isinstance(no, int):
                continue

            nisn = row[1]
            no_induk = row[2]
            nama = row[3]

            # Menentukan JK dari kolom L/P
            jk = ""

            if len(row) > 4:
                if row[4]:
                    jk = "L"

            if len(row) > 5:
                if row[5]:
                    jk = "P"

            if nama:

                data_siswa.append({

                    "nisn": str(nisn) if nisn else "",
                    "no_induk": str(no_induk) if no_induk else "",
                    "nama": str(nama).strip(),
                    "jk": jk,
                    "kelas": kelas_aktif

                })

        kosongkan_data_siswa()

        simpan_siswa(data_siswa)

        messagebox.showinfo(
            "Import Berhasil",
            f"{len(data_siswa)} siswa berhasil disimpan ke database."
        )

    except Exception as e:

        messagebox.showerror(
            "Error",
            str(e)
        )